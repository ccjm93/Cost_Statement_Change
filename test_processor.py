from __future__ import annotations

import tempfile
import unittest
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill

from excel_change_processor import (
    CHANGE_HEADER,
    CHANGED_LABEL,
    ORIGINAL_LABEL,
    list_sheet_names,
    preferred_sheet_name,
    process_workbook,
)


class ProcessorTest(unittest.TestCase):
    def test_process_sample_workbook(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            input_path = Path(tmp) / "sample.xlsx"
            output_path = Path(tmp) / "sample_설계변경자동화.xlsx"
            self._create_sample(input_path)

            result = process_workbook(input_path, "내역서", insert_before_col=2, output_path=output_path)

            self.assertEqual(result.processed_rows, 3)
            self.assertTrue(output_path.exists())

            workbook = load_workbook(output_path, data_only=False)
            try:
                ws = workbook["내역서"]
                self.assertEqual(ws["B2"].value, CHANGE_HEADER)
                self.assertEqual(ws["B3"].value, ORIGINAL_LABEL)
                self.assertEqual(ws["B4"].value, CHANGED_LABEL)
                self.assertEqual(ws["B5"].value, ORIGINAL_LABEL)
                self.assertEqual(ws["B6"].value, CHANGED_LABEL)
                self.assertEqual(ws["B7"].value, ORIGINAL_LABEL)
                self.assertEqual(ws["B8"].value, CHANGED_LABEL)
                self.assertEqual(ws["F3"].value, "=C3+D3+E3")
                self.assertEqual(ws["F5"].value, "=참조!B1*3")
                self.assertEqual(ws["C7"].value, "=SUM(C3,C5)")
                self.assertEqual(ws["F7"].value, "=SUM(F3,F5)")
                # 변경 행 금액: 같은 시트 참조만 있으면 당초 행 참조를 변경 행 참조로 옮긴다.
                self.assertEqual(ws["F4"].value, "=C4+D4+E4")
                self.assertEqual(ws["F8"].value, "=SUM(F4,F6)")
                # 변경 행 금액: 당초 수식이 다른 시트를 참조하면 노무비+재료비+경비로 재설정한다.
                self.assertEqual(ws["F6"].value, "=C6+D6+E6")
                # 변경 행: 같은 시트에서 당초 행을 가리키는 참조는 한 행 아래(변경 행)로 옮긴다.
                self.assertEqual(ws["C8"].value, "=SUM(C4,C6)")
                self.assertEqual(ws["D8"].value, "=SUM(D4,D6)")
                # 변경 행: 다른 시트 참조는 당초 행과 동일하게 유지한다.
                self.assertEqual(ws["G3"].value, "=참조!B1")
                self.assertEqual(ws["G4"].value, "=참조!B1")
                # 데이터 행 밖의 수식도 열/행 삽입에 맞게 보정된다. (원본 F6 -> G9)
                self.assertEqual(ws["G9"].value, "=F7")
                # 다른 시트에서 내역서를 참조하는 수식도 보정된다.
                self.assertEqual(workbook["참조"]["A1"].value, "=내역서!F3")
                self.assertEqual(ws["A1"].value, "공사 내역")
                self.assertIn("A1:G1", {str(range_) for range_ in ws.merged_cells.ranges})
                self.assertEqual(ws["A1"].fill.fgColor.rgb, "FF1F4E78")
                self.assertEqual(ws["C4"].font.color.rgb, "FFFF0000")
            finally:
                workbook.close()

    def test_workspace_sample_if_present(self) -> None:
        workspace = Path(__file__).resolve().parent
        samples = sorted(
            path
            for path in workspace.glob("sample_*.xlsx")
            if not path.name.endswith("_설계변경자동화.xlsx")
        )
        if not samples:
            self.skipTest("workspace sample workbook not found")

        sample = samples[0]
        with tempfile.TemporaryDirectory() as tmp:
            output_path = Path(tmp) / f"{sample.stem}_설계변경자동화.xlsx"
            sheet_name = preferred_sheet_name(list_sheet_names(sample))
            self.assertIsNotNone(sheet_name)

            result = process_workbook(
                sample,
                sheet_name,
                insert_before_col=1,
                output_path=output_path,
            )

            self.assertGreater(result.processed_rows, 0)
            self.assertTrue(output_path.exists())

    @staticmethod
    def _create_sample(path: Path) -> None:
        wb = Workbook()
        ws = wb.active
        ws.title = "내역서"
        ws.merge_cells("A1:F1")
        ws["A1"] = "공사 내역"
        ws["A1"].fill = PatternFill("solid", fgColor="FF1F4E78")
        ws["A1"].font = Font(color="FFFFFFFF", bold=True)
        headers = ["품명", "노무비", "재료비", "경비", "금액", "비고"]
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(2, col)
            cell.value = header
            cell.font = Font(bold=True)
        ws.append(["터파기", 100, 200, 30, "=B3+C3+D3", ""])
        ws.append(["되메우기", 120, 220, 40, "=B4+C4+D4", ""])
        ws.append(["소계", "=SUM(B3:B4)", "=SUM(C3:C4)", "=SUM(D3:D4)", "=SUM(E3:E4)", ""])
        ws["E4"] = "=참조!B1*3"  # 금액 열의 다른 시트 참조 수식
        ws["F3"] = "=참조!B1"  # 데이터 행의 다른 시트 참조 수식
        ws["F6"] = "=E5"  # 데이터 행 밖의 수식 (금액 열 수식 없음)
        ref_sheet = wb.create_sheet("참조")
        ref_sheet["A1"] = "=내역서!E3"
        ref_sheet["B1"] = 7
        wb.save(path)
        wb.close()


if __name__ == "__main__":
    unittest.main()
