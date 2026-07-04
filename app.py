from __future__ import annotations

import threading
import traceback
from pathlib import Path
from tkinter import BOTH, END, LEFT, RIGHT, Button, Frame, Label, Listbox, Tk, filedialog, messagebox
from tkinter import ttk

from openpyxl.utils import get_column_letter

from excel_change_processor import (
    default_output_path,
    inspect_header_columns,
    list_sheet_names,
    preferred_sheet_name,
    process_workbook,
)


class CostStatementApp:
    def __init__(self, root: Tk) -> None:
        self.root = root
        self.root.title("설계변경 내역서 자동화")
        self.root.geometry("760x520")
        self.file_path: Path | None = None
        self.sheet_names: list[str] = []
        self.detected_columns: dict[str, int] = {}
        self.busy = False

        self.file_label = Label(root, text="엑셀 파일을 선택하세요.", anchor="w")
        self.file_label.pack(fill="x", padx=16, pady=(16, 8))

        top = Frame(root)
        top.pack(fill="x", padx=16)
        self.file_button = Button(top, text="파일 선택", command=self.select_file)
        self.detect_button = Button(top, text="열 자동탐색", command=self.detect_columns)
        self.run_button = Button(top, text="자동화 실행", command=self.run_process)
        self.file_button.pack(side=LEFT)
        self.detect_button.pack(side=LEFT, padx=8)
        self.run_button.pack(side=RIGHT)
        self.buttons = [self.file_button, self.detect_button, self.run_button]

        body = Frame(root)
        body.pack(fill=BOTH, expand=True, padx=16, pady=12)

        sheet_frame = Frame(body)
        sheet_frame.pack(side=LEFT, fill=BOTH, expand=True)
        Label(sheet_frame, text="처리할 시트").pack(anchor="w")
        self.sheet_list = Listbox(sheet_frame, exportselection=False, height=10)
        self.sheet_list.pack(fill=BOTH, expand=True)

        col_frame = Frame(body)
        col_frame.pack(side=RIGHT, fill=BOTH, expand=True, padx=(16, 0))
        Label(col_frame, text="당초/변경 열을 삽입할 기준 열").pack(anchor="w")
        self.column_list = Listbox(col_frame, exportselection=False, height=10)
        self.column_list.pack(fill=BOTH, expand=True)

        self.status = Listbox(root, height=9)
        self.status.pack(fill=BOTH, padx=16, pady=(0, 16))
        self.progress = ttk.Progressbar(root, mode="indeterminate")
        self.progress.pack(fill="x", padx=16, pady=(0, 12))
        self.log("대기 중")

    def select_file(self) -> None:
        if self.busy:
            return
        selected = filedialog.askopenfilename(
            title="설계변경 내역서 엑셀 선택",
            filetypes=[("Excel workbook", "*.xlsx")],
        )
        if not selected:
            return
        self.file_path = Path(selected)
        self.file_label.config(text=str(self.file_path))
        self.sheet_names = []
        self.detected_columns = {}
        self.sheet_list.delete(0, END)
        self.log(f"파일 선택: {self.file_path.name}")
        self.log("시트 목록을 읽는 중...")
        self.set_busy(True)
        threading.Thread(target=self._load_file_worker, args=(self.file_path,), daemon=True).start()

    def _load_file_worker(self, file_path: Path) -> None:
        try:
            sheet_names = list_sheet_names(file_path)
        except Exception as exc:
            detail = traceback.format_exc()
            self.root.after(0, self._file_load_failed, exc, detail)
            return
        self.root.after(0, self._file_load_done, file_path, sheet_names)

    def _file_load_done(self, file_path: Path, sheet_names: list[str]) -> None:
        self.set_busy(False)
        if self.file_path != file_path:
            return
        self.sheet_names = sheet_names
        self.sheet_list.delete(0, END)
        for sheet_name in self.sheet_names:
            self.sheet_list.insert(END, sheet_name)
        preferred = preferred_sheet_name(self.sheet_names)
        if preferred:
            index = self.sheet_names.index(preferred)
            self.sheet_list.selection_set(index)
            self.sheet_list.activate(index)
        self.log(f"시트 {len(sheet_names)}개를 읽었습니다.")
        self.detect_columns()

    def _file_load_failed(self, exc: Exception, detail: str) -> None:
        self.set_busy(False)
        self.log(f"파일 읽기 실패: {exc}")
        self.log(detail.splitlines()[-1] if detail else str(exc))
        messagebox.showerror("파일 읽기 실패", str(exc))

    def detect_columns(self) -> None:
        if self.busy:
            return
        if not self.file_path:
            messagebox.showinfo("안내", "먼저 엑셀 파일을 선택하세요.")
            return
        sheet_name = self.selected_sheet()
        if not sheet_name:
            messagebox.showinfo("안내", "처리할 시트를 선택하세요.")
            return
        self.log("열 자동탐색 중...")
        self.set_busy(True)
        threading.Thread(
            target=self._detect_columns_worker,
            args=(self.file_path, sheet_name),
            daemon=True,
        ).start()

    def _detect_columns_worker(self, file_path: Path, sheet_name: str) -> None:
        try:
            header_row, columns = inspect_header_columns(file_path, sheet_name)
        except Exception as exc:
            detail = traceback.format_exc()
            self.root.after(0, self._detect_columns_failed, exc, detail)
            return
        self.root.after(0, self._detect_columns_done, file_path, sheet_name, header_row, columns)

    def _detect_columns_done(
        self,
        file_path: Path,
        sheet_name: str,
        header_row: int,
        columns: dict[str, int],
    ) -> None:
        self.set_busy(False)
        if self.file_path != file_path or self.selected_sheet() != sheet_name:
            return
        self.detected_columns = columns
        self.column_list.delete(0, END)
        max_col = max(columns.values()) + 1
        for col in range(1, max_col + 1):
            letter = get_column_letter(col)
            self.column_list.insert(END, f"{letter}열 앞")
        default_col = min(columns.values())
        self.column_list.selection_set(default_col - 1)
        self.column_list.activate(default_col - 1)
        self.log(
            "열 자동탐색 완료: "
            f"헤더 {header_row}행, "
            f"노무비 {get_column_letter(columns['labor'])}, "
            f"재료비 {get_column_letter(columns['material'])}, "
            f"경비 {get_column_letter(columns['expense'])}, "
            f"금액 {get_column_letter(columns['amount'])}"
        )

    def _detect_columns_failed(self, exc: Exception, detail: str) -> None:
        self.set_busy(False)
        self.log(f"열 자동탐색 실패: {exc}")
        self.log(detail.splitlines()[-1] if detail else str(exc))
        messagebox.showerror("열 자동탐색 실패", str(exc))

    def run_process(self) -> None:
        if self.busy:
            return
        if not self.file_path:
            messagebox.showinfo("안내", "먼저 엑셀 파일을 선택하세요.")
            return
        sheet_name = self.selected_sheet()
        insert_col = self.selected_insert_col()
        if not sheet_name or not insert_col:
            messagebox.showinfo("안내", "시트와 삽입 기준 열을 선택하세요.")
            return

        output_path = default_output_path(self.file_path)
        self.log("자동화 실행 중...")
        self.set_busy(True)
        thread = threading.Thread(
            target=self._run_process_worker,
            args=(sheet_name, insert_col, output_path),
            daemon=True,
        )
        thread.start()

    def _run_process_worker(self, sheet_name: str, insert_col: int, output_path: Path) -> None:
        try:
            result = process_workbook(
                input_path=self.file_path,
                sheet_name=sheet_name,
                insert_before_col=insert_col,
                output_path=output_path,
            )
        except Exception as exc:
            detail = traceback.format_exc()
            self.root.after(0, self._process_failed, exc, detail)
            return
        self.root.after(0, self._process_done, result)

    def _process_done(self, result) -> None:
        self.set_busy(False)
        self.log(f"완료: {result.processed_rows}개 데이터 행 처리")
        self.log(f"저장 위치: {result.output_path}")
        self.log(", ".join(result.validation))
        messagebox.showinfo("완료", f"처리가 완료되었습니다.\n{result.output_path}")

    def _process_failed(self, exc: Exception, detail: str) -> None:
        self.set_busy(False)
        self.log(f"실패: {exc}")
        self.log(detail.splitlines()[-1] if detail else str(exc))
        messagebox.showerror("처리 실패", str(exc))

    def selected_sheet(self) -> str | None:
        selection = self.sheet_list.curselection()
        if not selection:
            return None
        return self.sheet_list.get(selection[0])

    def selected_insert_col(self) -> int | None:
        selection = self.column_list.curselection()
        if not selection:
            return None
        return int(selection[0]) + 1

    def log(self, message: str) -> None:
        self.status.insert(END, message)
        self.status.see(END)

    def set_busy(self, busy: bool) -> None:
        self.busy = busy
        state = "disabled" if busy else "normal"
        for button in self.buttons:
            button.configure(state=state)
        if busy:
            self.progress.start(12)
        else:
            self.progress.stop()
        self.root.title(f"설계변경 내역서 자동화{' - 처리 중' if busy else ''}")


def main() -> None:
    root = Tk()
    CostStatementApp(root)
    root.mainloop()


if __name__ == "__main__":
    main()
