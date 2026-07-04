from __future__ import annotations

from copy import copy
from dataclasses import dataclass
from pathlib import Path
import re
import tempfile
from typing import Iterable
import zipfile
import xml.etree.ElementTree as ET

from openpyxl import load_workbook
from openpyxl.cell.cell import Cell, MergedCell
from openpyxl.styles import Font
from openpyxl.utils import column_index_from_string, get_column_letter


CHANGE_HEADER = "당초/변경"
ORIGINAL_LABEL = "당초"
CHANGED_LABEL = "변경"
REQUIRED_HEADERS = {
    "labor": "노무비",
    "material": "재료비",
    "expense": "경비",
    "amount": "금액",
}
RED_RGB = "FFFF0000"


@dataclass(frozen=True)
class ProcessResult:
    output_path: Path
    sheet_name: str
    header_row: int
    insert_col: int
    processed_rows: int
    detected_columns: dict[str, int]
    validation: list[str]


def default_output_path(input_path: str | Path) -> Path:
    path = Path(input_path)
    return path.with_name(f"{path.stem}_설계변경자동화{path.suffix}")


def list_sheet_names(input_path: str | Path) -> list[str]:
    workbook, cleanup_path = _load_workbook_resilient(input_path, read_only=True, data_only=False)
    try:
        return list(workbook.sheetnames)
    finally:
        workbook.close()
        _cleanup_temp_file(cleanup_path)


def preferred_sheet_name(sheet_names: Iterable[str]) -> str | None:
    names = list(sheet_names)
    for target in ("내역서", "내역"):
        if target in names:
            return target
    return names[0] if names else None


def inspect_header_columns(input_path: str | Path, sheet_name: str) -> tuple[int, dict[str, int]]:
    workbook, cleanup_path = _load_workbook_resilient(input_path, read_only=False, data_only=False)
    try:
        if sheet_name not in workbook.sheetnames:
            raise ValueError(f"시트를 찾을 수 없습니다: {sheet_name}")
        return find_header_columns(workbook[sheet_name])
    finally:
        workbook.close()
        _cleanup_temp_file(cleanup_path)


def process_workbook(
    input_path: str | Path,
    sheet_name: str,
    insert_before_col: int,
    output_path: str | Path | None = None,
) -> ProcessResult:
    input_path = Path(input_path)
    if input_path.suffix.lower() != ".xlsx":
        raise ValueError("v1은 .xlsx 파일만 지원합니다.")
    if insert_before_col < 1:
        raise ValueError("삽입 열 번호는 1 이상이어야 합니다.")

    output = Path(output_path) if output_path else default_output_path(input_path)
    workbook, cleanup_path = _load_workbook_resilient(input_path)
    try:
        if sheet_name not in workbook.sheetnames:
            raise ValueError(f"시트를 찾을 수 없습니다: {sheet_name}")
        worksheet = workbook[sheet_name]

        original_merged_ranges = [str(range_) for range_ in worksheet.merged_cells.ranges]
        header_row, detected_columns = find_header_columns(worksheet)
        if insert_before_col > worksheet.max_column + 1:
            raise ValueError("선택한 삽입 위치가 시트 범위를 벗어났습니다.")
        data_rows = _formula_rows(worksheet, detected_columns["amount"], header_row)
        original_formulas = _snapshot_all_formulas(worksheet)
        ref_mapper = _FormulaRefMapper(data_rows, insert_before_col)

        _unmerge_ranges(worksheet, original_merged_ranges)
        worksheet.insert_cols(insert_before_col)
        _restore_merged_ranges_after_col_insert(
            worksheet,
            original_merged_ranges,
            insert_before_col,
        )
        _copy_inserted_column_style(worksheet, insert_before_col, header_row)
        worksheet.cell(header_row, insert_before_col).value = CHANGE_HEADER

        shifted_columns = {
            key: col + 1 if col >= insert_before_col else col
            for key, col in detected_columns.items()
        }

        processed_original_rows: list[int] = []
        offset = 0
        for original_row in data_rows:
            row_idx = original_row + offset
            change_row_idx = row_idx + 1
            worksheet.insert_rows(change_row_idx)
            _copy_row(worksheet, row_idx, change_row_idx)

            worksheet.cell(row_idx, insert_before_col).value = ORIGINAL_LABEL
            worksheet.cell(change_row_idx, insert_before_col).value = CHANGED_LABEL
            _apply_red_font(worksheet, change_row_idx)

            processed_original_rows.append(row_idx)
            offset += 1

        data_row_set = set(data_rows)
        original_positions = {ref_mapper.map_row(row) for row in data_rows}
        for (row, col), formula in original_formulas.items():
            target_row = ref_mapper.map_row(row)
            target_col = ref_mapper.map_col(col)
            rewritten = _rewrite_same_sheet_formula(formula, ref_mapper)
            worksheet.cell(target_row, target_col).value = rewritten
            if row in data_row_set:
                # 변경 행 수식: 다른 시트 참조는 당초 행과 동일하게 두고,
                # 같은 시트에서 당초 행을 가리키는 참조만 한 행 아래(변경 행)로 옮긴다.
                worksheet.cell(target_row + 1, target_col).value = _shift_change_row_formula(
                    rewritten, original_positions
                )

        for row_idx in processed_original_rows:
            amount_value = worksheet.cell(row_idx, shifted_columns["amount"]).value
            if _amount_formula_needs_reset(amount_value):
                _set_amount_formula(worksheet, row_idx + 1, shifted_columns)

        _rewrite_cross_sheet_references(workbook, sheet_name, ref_mapper)

        validation = validate_workbook(
            worksheet=worksheet,
            header_row=header_row,
            insert_col=insert_before_col,
            columns=shifted_columns,
            original_rows=processed_original_rows,
            original_merged_ranges=original_merged_ranges,
        )
        if validation:
            raise ValueError("검증 실패: " + " / ".join(validation))

        _force_recalculate_on_open(workbook)
        output.parent.mkdir(parents=True, exist_ok=True)
        workbook.save(output)
    finally:
        workbook.close()
        _cleanup_temp_file(cleanup_path)

    return ProcessResult(
        output_path=output,
        sheet_name=sheet_name,
        header_row=header_row,
        insert_col=insert_before_col,
        processed_rows=len(processed_original_rows),
        detected_columns=shifted_columns,
        validation=["검증 통과"],
    )


def find_header_columns(worksheet, max_scan_rows: int = 40) -> tuple[int, dict[str, int]]:
    layered = _find_layered_header_columns(worksheet, max_scan_rows)
    if layered:
        return layered

    for row in range(1, min(worksheet.max_row, max_scan_rows) + 1):
        found: dict[str, int] = {}
        for cell in worksheet[row]:
            text = _normalized_text(cell.value)
            if not text:
                continue
            for key, header in REQUIRED_HEADERS.items():
                if key not in found and header in text:
                    found[key] = cell.column
        if set(found) == set(REQUIRED_HEADERS):
            return row, found

    missing = ", ".join(REQUIRED_HEADERS.values())
    raise ValueError(f"헤더 행을 찾지 못했습니다. 필요한 헤더: {missing}")


def _find_layered_header_columns(worksheet, max_scan_rows: int) -> tuple[int, dict[str, int]] | None:
    for row in range(1, min(worksheet.max_row, max_scan_rows)):
        found: dict[str, int] = {}
        for col in range(1, worksheet.max_column + 1):
            group_text = _normalized_text(_effective_cell_value(worksheet, row, col))
            sub_text = _normalized_text(_effective_cell_value(worksheet, row + 1, col))
            if "금액" not in sub_text:
                continue
            if "계" in group_text and "amount" not in found:
                found["amount"] = col
            elif "노무비" in group_text and "labor" not in found:
                found["labor"] = col
            elif "재료비" in group_text and "material" not in found:
                found["material"] = col
            elif "경비" in group_text and "expense" not in found:
                found["expense"] = col
        if set(found) == set(REQUIRED_HEADERS):
            return row, found
    return None


def _effective_cell_value(worksheet, row: int, col: int):
    cell = worksheet.cell(row, col)
    if cell.value is not None:
        return cell.value
    for merged_range in worksheet.merged_cells.ranges:
        if cell.coordinate in merged_range:
            return worksheet.cell(merged_range.min_row, merged_range.min_col).value
    return None


def validate_workbook(
    worksheet,
    header_row: int,
    insert_col: int,
    columns: dict[str, int],
    original_rows: Iterable[int],
    original_merged_ranges: Iterable[str],
) -> list[str]:
    errors: list[str] = []
    original_rows = list(original_rows)
    original_positions = set(original_rows)
    if worksheet.cell(header_row, insert_col).value != CHANGE_HEADER:
        errors.append("당초/변경 헤더가 없습니다.")

    for row_idx in original_rows:
        change_row_idx = row_idx + 1
        if worksheet.cell(row_idx, insert_col).value != ORIGINAL_LABEL:
            errors.append(f"{row_idx}행 당초 표시 누락")
        if worksheet.cell(change_row_idx, insert_col).value != CHANGED_LABEL:
            errors.append(f"{change_row_idx}행 변경 표시 누락")
        if not _row_has_red_font(worksheet, change_row_idx):
            errors.append(f"{change_row_idx}행 빨간 글자 서식 누락")

        amount_value = worksheet.cell(row_idx, columns["amount"]).value
        if _amount_formula_needs_reset(amount_value):
            expected = _amount_formula(change_row_idx, columns)
        else:
            expected = _shift_change_row_formula(amount_value, original_positions)
        if worksheet.cell(change_row_idx, columns["amount"]).value != expected:
            errors.append(f"{change_row_idx}행 금액 수식 오류")

        for col in range(1, worksheet.max_column + 1):
            if col in (insert_col, columns["amount"]):
                continue
            original_value = worksheet.cell(row_idx, col).value
            if not (isinstance(original_value, str) and original_value.startswith("=")):
                continue
            expected_formula = _shift_change_row_formula(original_value, original_positions)
            if worksheet.cell(change_row_idx, col).value != expected_formula:
                errors.append(f"{change_row_idx}행 {get_column_letter(col)}열 수식이 변경 행 규칙과 다릅니다")

    shifted_ranges = set(_shift_merged_ranges_after_col_insert(original_merged_ranges, insert_col))
    current_ranges = {str(range_) for range_ in worksheet.merged_cells.ranges}
    if not shifted_ranges.issubset(current_ranges):
        errors.append("기존 병합 셀 범위가 일부 유지되지 않았습니다.")

    return errors


def _copy_inserted_column_style(worksheet, insert_col: int, header_row: int) -> None:
    source_col = insert_col + 1 if insert_col < worksheet.max_column else insert_col - 1
    if source_col < 1:
        return
    worksheet.column_dimensions[get_column_letter(insert_col)].width = worksheet.column_dimensions[
        get_column_letter(source_col)
    ].width
    max_row = max(worksheet.max_row, header_row)
    for row in range(1, max_row + 1):
        source = worksheet.cell(row, source_col)
        target = worksheet.cell(row, insert_col)
        _copy_cell_style(source, target)


def _load_workbook_resilient(input_path: str | Path, **kwargs):
    try:
        return load_workbook(input_path, **kwargs), None
    except ValueError as exc:
        if "could not assign names" not in str(exc):
            raise
        repaired_path = _repair_invalid_defined_names(input_path)
        return load_workbook(repaired_path, **kwargs), repaired_path


def _repair_invalid_defined_names(input_path: str | Path) -> Path:
    source = Path(input_path)
    target = Path(tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False).name)
    workbook_xml = "xl/workbook.xml"
    namespace = {"main": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}
    ET.register_namespace("", namespace["main"])

    with zipfile.ZipFile(source, "r") as zin, zipfile.ZipFile(target, "w", zipfile.ZIP_DEFLATED) as zout:
        for item in zin.infolist():
            data = zin.read(item.filename)
            if item.filename == workbook_xml:
                root = ET.fromstring(data)
                defined_names = root.find("main:definedNames", namespace)
                if defined_names is not None:
                    for defined_name in list(defined_names):
                        value = (defined_name.text or "").strip()
                        name = defined_name.attrib.get("name", "")
                        if value == "#N/A" or (name == "_xlnm.Print_Titles" and "!" not in value):
                            defined_names.remove(defined_name)
                    if len(defined_names) == 0:
                        root.remove(defined_names)
                data = ET.tostring(root, encoding="utf-8", xml_declaration=True)
            zout.writestr(item, data)
    return target


def _cleanup_temp_file(path: Path | None) -> None:
    if not path:
        return
    try:
        path.unlink(missing_ok=True)
    except OSError:
        pass


def _force_recalculate_on_open(workbook) -> None:
    workbook.calculation.fullCalcOnLoad = True
    workbook.calculation.forceFullCalc = True
    workbook.calculation.calcMode = "auto"


def _restore_merged_ranges_after_col_insert(
    worksheet,
    original_merged_ranges: Iterable[str],
    insert_col: int,
) -> None:
    for range_text in _shift_merged_ranges_after_col_insert(original_merged_ranges, insert_col):
        worksheet.merge_cells(range_text)


def _unmerge_ranges(worksheet, ranges: Iterable[str]) -> None:
    for range_text in ranges:
        worksheet.unmerge_cells(range_text)


def _formula_rows(worksheet, amount_col: int, header_row: int) -> list[int]:
    rows: list[int] = []
    for row in range(header_row + 1, worksheet.max_row + 1):
        value = worksheet.cell(row, amount_col).value
        if isinstance(value, str) and value.startswith("="):
            rows.append(row)
    return rows


def _snapshot_all_formulas(worksheet) -> dict[tuple[int, int], str]:
    formulas: dict[tuple[int, int], str] = {}
    for row in worksheet.iter_rows():
        for cell in row:
            value = cell.value
            if isinstance(value, str) and value.startswith("="):
                formulas[(cell.row, cell.column)] = value
    return formulas


class _FormulaRefMapper:
    def __init__(self, inserted_after_rows: Iterable[int], inserted_before_col: int) -> None:
        self.inserted_after_rows = sorted(inserted_after_rows)
        self.inserted_after_set = set(self.inserted_after_rows)
        self.inserted_before_col = inserted_before_col

    def map_row(self, row: int) -> int:
        return row + sum(1 for inserted_row in self.inserted_after_rows if inserted_row < row)

    def map_col(self, col: int) -> int:
        return col + 1 if col >= self.inserted_before_col else col

    def map_col_token(self, col_token: str) -> str:
        prefix = "$" if col_token.startswith("$") else ""
        col_index = column_index_from_string(col_token.replace("$", ""))
        return f"{prefix}{get_column_letter(self.map_col(col_index))}"

    def range_has_inserted_rows(self, start_row: int, end_row: int) -> bool:
        low, high = sorted((start_row, end_row))
        return any(low <= inserted_row <= high for inserted_row in self.inserted_after_set)


_CELL_TOKEN = r"(\$?[A-Z]{1,3})(\$?)(\d+)"
_RANGE_RE = re.compile(rf"(?<!!){_CELL_TOKEN}:{_CELL_TOKEN}")
_CELL_RE = re.compile(rf"(?<!!){_CELL_TOKEN}")
_SHEET_REF_RE = re.compile(
    r"(?:'[^']+'|[A-Za-z0-9_\.\-ㄱ-힝]+)!\$?[A-Z]{1,3}\$?\d+(?::\$?[A-Z]{1,3}\$?\d+)?"
)


def _rewrite_same_sheet_formula(formula: str, ref_mapper: _FormulaRefMapper) -> str:
    protected: list[str] = []
    rewritten_ranges: list[str] = []

    def protect(match: re.Match[str]) -> str:
        protected.append(match.group(0))
        return f"__SHEET_REF_{len(protected) - 1}__"

    def protect_rewritten_range(value: str) -> str:
        rewritten_ranges.append(value)
        return f"__RANGE_REF_{len(rewritten_ranges) - 1}__"

    formula = _SHEET_REF_RE.sub(protect, formula)

    def replace_range(match: re.Match[str]) -> str:
        start_col, start_abs, start_row_text, end_col, end_abs, end_row_text = match.groups()
        start_row = int(start_row_text)
        end_row = int(end_row_text)
        mapped_start_col = ref_mapper.map_col_token(start_col)
        mapped_end_col = ref_mapper.map_col_token(end_col)
        if start_col.replace("$", "") != end_col.replace("$", "") or not ref_mapper.range_has_inserted_rows(start_row, end_row):
            return (
                f"{mapped_start_col}{start_abs}{ref_mapper.map_row(start_row)}:"
                f"{mapped_end_col}{end_abs}{ref_mapper.map_row(end_row)}"
            )

        step = 1 if end_row >= start_row else -1
        refs = [
            f"{mapped_start_col}{start_abs}{ref_mapper.map_row(row)}"
            for row in range(start_row, end_row + step, step)
        ]
        return protect_rewritten_range(",".join(refs))

    formula = _RANGE_RE.sub(replace_range, formula)

    def replace_cell(match: re.Match[str]) -> str:
        col, row_abs, row_text = match.groups()
        return f"{ref_mapper.map_col_token(col)}{row_abs}{ref_mapper.map_row(int(row_text))}"

    formula = _CELL_RE.sub(replace_cell, formula)

    for index, value in enumerate(rewritten_ranges):
        formula = formula.replace(f"__RANGE_REF_{index}__", value)
    for index, value in enumerate(protected):
        formula = formula.replace(f"__SHEET_REF_{index}__", value)
    return formula


def _shift_change_row_formula(formula: str, original_row_positions: set[int]) -> str:
    """변경 행 수식을 만든다.

    - 다른 시트 참조: 당초 행과 동일하게 유지한다.
    - 같은 시트 참조: 당초 행(변경 행 짝이 있는 행)을 가리키는 참조만 한 행 아래로 옮긴다.
      (당초 행이 'A5'를 참조하면 변경 행은 'A6'을 참조)
    """
    protected: list[str] = []

    def protect(match: re.Match[str]) -> str:
        protected.append(match.group(0))
        return f"__SHEET_REF_{len(protected) - 1}__"

    result = _SHEET_REF_RE.sub(protect, formula)

    def shift(match: re.Match[str]) -> str:
        col_token, row_abs, row_text = match.groups()
        row = int(row_text)
        if row in original_row_positions:
            row += 1
        return f"{col_token}{row_abs}{row}"

    result = _CELL_RE.sub(shift, result)

    for index, value in enumerate(protected):
        result = result.replace(f"__SHEET_REF_{index}__", value)
    return result


def _rewrite_cross_sheet_references(workbook, sheet_name: str, ref_mapper: _FormulaRefMapper) -> None:
    """다른 시트의 수식이 처리 대상 시트를 참조하면 열/행 삽입에 맞게 참조를 보정한다."""
    quoted = re.escape(f"'{sheet_name}'")
    bare = re.escape(sheet_name)
    pattern = re.compile(
        rf"(?<![\w.ㄱ-힣'])({quoted}|{bare})!(\$?[A-Z]{{1,3}}\$?\d+)(?::(\$?[A-Z]{{1,3}}\$?\d+))?"
    )
    cell_re = re.compile(r"(\$?[A-Z]{1,3})(\$?)(\d+)")

    def map_ref(ref: str) -> str:
        col_token, row_abs, row_text = cell_re.fullmatch(ref).groups()
        return f"{ref_mapper.map_col_token(col_token)}{row_abs}{ref_mapper.map_row(int(row_text))}"

    def replace(match: re.Match[str]) -> str:
        sheet_part, start, end = match.groups()
        mapped = f"{sheet_part}!{map_ref(start)}"
        if end:
            mapped += f":{map_ref(end)}"
        return mapped

    for name in workbook.sheetnames:
        if name == sheet_name:
            continue
        other = workbook[name]
        for row in other.iter_rows():
            for cell in row:
                value = cell.value
                if isinstance(value, str) and value.startswith("=") and sheet_name in value:
                    rewritten = pattern.sub(replace, value)
                    if rewritten != value:
                        cell.value = rewritten


def _copy_row(worksheet, source_row: int, target_row: int) -> None:
    worksheet.row_dimensions[target_row].height = worksheet.row_dimensions[source_row].height
    for col in range(1, worksheet.max_column + 1):
        source = worksheet.cell(source_row, col)
        target = worksheet.cell(target_row, col)
        if isinstance(source, MergedCell):
            continue
        target.value = source.value
        _copy_cell_style(source, target)
        if source.hyperlink:
            target._hyperlink = copy(source.hyperlink)
        if source.comment:
            target.comment = copy(source.comment)


def _copy_cell_style(source: Cell, target: Cell) -> None:
    if not source.has_style:
        return
    target._style = copy(source._style)
    target.number_format = source.number_format
    target.protection = copy(source.protection)
    target.alignment = copy(source.alignment)


def _apply_red_font(worksheet, row_idx: int) -> None:
    for col in range(1, worksheet.max_column + 1):
        cell = worksheet.cell(row_idx, col)
        if isinstance(cell, MergedCell):
            continue
        cell.font = _red_font(cell.font)


def _red_font(font: Font) -> Font:
    new_font = copy(font)
    new_font.color = RED_RGB
    return new_font


def _row_has_red_font(worksheet, row_idx: int) -> bool:
    for col in range(1, worksheet.max_column + 1):
        cell = worksheet.cell(row_idx, col)
        if isinstance(cell, MergedCell) or cell.value is None:
            continue
        color = cell.font.color
        if color is None or color.type != "rgb" or color.rgb != RED_RGB:
            return False
    return True


def _amount_formula_needs_reset(formula) -> bool:
    """당초 금액 수식이 다른 시트를 참조할 때만 변경 행 금액 수식을 노무비+재료비+경비로 재설정한다."""
    return isinstance(formula, str) and bool(_SHEET_REF_RE.search(formula))


def _set_amount_formula(worksheet, row_idx: int, columns: dict[str, int]) -> None:
    worksheet.cell(row_idx, columns["amount"]).value = _amount_formula(row_idx, columns)


def _amount_formula(row_idx: int, columns: dict[str, int]) -> str:
    labor = f"{get_column_letter(columns['labor'])}{row_idx}"
    material = f"{get_column_letter(columns['material'])}{row_idx}"
    expense = f"{get_column_letter(columns['expense'])}{row_idx}"
    return f"={labor}+{material}+{expense}"


def _shift_merged_ranges_after_col_insert(ranges: Iterable[str], insert_col: int) -> list[str]:
    shifted: list[str] = []
    for range_text in ranges:
        start, end = range_text.split(":") if ":" in range_text else (range_text, range_text)
        start_col, start_row = _split_coordinate(start)
        end_col, end_row = _split_coordinate(end)
        if start_col >= insert_col:
            start_col += 1
            end_col += 1
        elif start_col < insert_col <= end_col:
            end_col += 1
        shifted.append(
            f"{get_column_letter(start_col)}{start_row}:{get_column_letter(end_col)}{end_row}"
        )
    return shifted


def _split_coordinate(coordinate: str) -> tuple[int, int]:
    letters = "".join(ch for ch in coordinate if ch.isalpha())
    digits = "".join(ch for ch in coordinate if ch.isdigit())
    col = 0
    for char in letters.upper():
        col = col * 26 + (ord(char) - ord("A") + 1)
    return col, int(digits)


def _normalized_text(value) -> str:
    if value is None:
        return ""
    return str(value).replace(" ", "").replace("\n", "").strip()
